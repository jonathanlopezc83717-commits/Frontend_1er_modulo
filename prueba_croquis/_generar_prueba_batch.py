"""Prueba del flujo de gui_batch() sin ventanas.

Crea una carpeta raiz con CAD + 2 subcarpetas con Excel valido + 1 subcarpeta
sin Excel (debe ser ignorada). Replica la logica de gui_batch() sin el
multiselect: procesa todos los candidatos. Si pasa, --batch real (con ventanas)
funciona.
"""
import importlib.util
import pathlib
import re

import ezdxf
import openpyxl

RAIZ = pathlib.Path(__file__).resolve().parent
SCRIPTS = RAIZ.parent / "scripts"

spec = importlib.util.spec_from_file_location(
    "dwg_to_croquis", SCRIPTS / "dwg-to-croquis.py")
mod = importlib.util.module_from_spec(spec)
spec.loader.exec_module(mod)

carpeta_raiz = RAIZ / "datos_batch"
if carpeta_raiz.exists():
    import shutil
    shutil.rmtree(carpeta_raiz)
carpeta_raiz.mkdir()

# CAD de juguete compartido
cx, cy = 1000.0, 1000.0
doc = ezdxf.new(setup=True)
msp = doc.modelspace()
for dx, dy in [(-400, 0), (0, -400)]:
    msp.add_line((cx + dx, cy + dy), (cx - dx, cy - dy))
msp.add_circle((cx, cy), radius=20)
doc.saveas(str(carpeta_raiz / "plano.dxf"))

# subcarpetas con Excel valido
puntos = [("01_PuntoA", 1000.0, 1000.0),
          ("02_PuntoB", 950.0, 1050.0)]
for nombre, x, y in puntos:
    sub = carpeta_raiz / nombre
    sub.mkdir()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['B1'] = x
    ws['C1'] = y
    wb.save(str(sub / f"{nombre}_20260714.xlsx"))

# subcarpeta basura (sin Excel valido) -> debe ser ignorada
(carpeta_raiz / "99_basura").mkdir()
(carpeta_raiz / "99_basura" / "notas.txt").write_text("x")

# --- replica de gui_batch() sin multiselect ---
PATRON = re.compile(r'^\d+_.+_.+\.xlsx$', re.IGNORECASE)
cad = sorted(carpeta_raiz.glob("*.dwg")) + sorted(carpeta_raiz.glob("*.dxf"))
assert len(cad) == 1, f"CAD: {len(cad)}"
archivo_cad = str(cad[0])

candidatos = []
for sub in sorted(p for p in carpeta_raiz.iterdir() if p.is_dir()):
    xl = sorted(f for f in sub.iterdir() if PATRON.match(f.name))
    if xl:
        candidatos.append((sub, xl[0]))
assert len(candidatos) == 2, f"Debia 2 candidatos, hay {len(candidatos)}: {candidatos}"

SIZE = 100.0
ok_list, fallos = [], []
for sub, xl in candidatos:
    try:
        wb = openpyxl.load_workbook(str(xl), data_only=True, read_only=True)
        ws = wb.active
        bx, cyv = ws['B1'].value, ws['C1'].value
        wb.close()
        out = str(sub / (sub.name + "_croquis.png"))
        mod.renderizar(archivo_cad, bx, cyv, SIZE, out, dpi=150,
                       anclar_esquina=False)
        kb = pathlib.Path(out).stat().st_size / 1024
        ok_list.append(f"{sub.name}: {kb:.1f}KB en ({bx},{cyv}) -> {out}")
    except Exception as e:
        fallos.append(f"{sub.name}: {e}")

print(f"CAD compartido: {archivo_cad}")
print(f"Candidatos (subcarpetas con Excel): {len(candidatos)}")
print(f"OK {len(ok_list)}/{len(candidatos)}")
for o in ok_list:
    print("  +", o)
for f in fallos:
    print("  -", f)
