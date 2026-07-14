#!/usr/bin/env python3
"""DWG/DXF -> PNG (croquis de localizacion).

Renderiza una ventana de `size` x `size` cm (unidades del dibujo) anclada en
las coordenadas X,Y del primer punto, y entrega un PNG listo para subir al
slot "Croquis de localizacion" del modulo Formato.

Pipeline:
  DWG --(dwg2dxf de LibreDWG)--> DXF --(ezdxf + matplotlib)--> PNG
  DXF se procesa directo (AutoCAD "Guardar como DXF" tambien sirve).

Uso:
  python scripts/dwg-to-croquis.py --input pl.dxf --x 1042 --y 3318 \
      --size 100 --out croquis.png

Las unidades del DWG ya son cm, asi que `--size 100` = 100x100 cm y `--x/--y`
van en las mismas unidades que trae el modulo de sincronizacion.
"""
import argparse
import os
import shutil
import subprocess
import sys
import tempfile

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import ezdxf
from ezdxf.addons.drawing import RenderContext, Frontend
from ezdxf.addons.drawing.matplotlib import MatplotlibBackend

CM_POR_PULGADA = 2.54


def _com_busy(fn, intentos=120, espera=1.0):
    """Reintenta llamadas COM si AutoCAD esta ocupado (RPC_E_CALL_REJECTED /
    RETRYLATER). Budget ~2 min. Patron de server/croquis_com.py."""
    import time
    import pywintypes
    _BUSY = {-2147418111, -2147418102}
    for k in range(intentos):
        try:
            return fn()
        except pywintypes.com_error as e:
            if getattr(e, "hresult", None) in _BUSY and k + 1 < intentos:
                time.sleep(espera)
                continue
            raise
    raise pywintypes.com_error(-2147418111, "RPC_E_CALL_REJECTED", None, None)


def dwg_a_dxf(dwg_path: str) -> str:
    """Convierte DWG->DXF. Preferencia: dwg2dxf (LibreDWG). Si no esta en PATH,
    fallback a la instancia de AutoCAD ya abierta via COM (SaveAs acR12_dxf=1,
    universalmente legible por ezdxf). Aborta con instrucciones si ambos fallan."""
    exe = shutil.which("dwg2dxf")
    if exe:
        tmp = tempfile.NamedTemporaryFile(suffix=".dxf", delete=False).name
        subprocess.run([exe, "-y", "-o", tmp, dwg_path], check=True,
                       capture_output=True)
        return tmp
    # Fallback COM: adhesivo a la instancia abierta de AutoCAD. Reusa pywin32
    # (ya en el venv por server/croquis_com.py). Si AutoCAD no corre, falla limpio.
    try:
        import win32com.client
        import pythoncom
        pythoncom.CoInitialize()
        acad = _com_busy(lambda: win32com.client.GetActiveObject("AutoCAD.Application"))
        tmp = tempfile.NamedTemporaryFile(suffix=".dxf", delete=False).name
        docs = acad.Documents
        nombre = os.path.basename(dwg_path).lower()
        doc = None
        for i in range(_com_busy(lambda: docs.Count)):
            d = _com_busy(lambda i=i: docs.Item(i))
            if os.path.basename(str(_com_busy(lambda d=d: d.FullName))).lower() == nombre:
                doc = d
                break
        if doc is None:
            doc = _com_busy(lambda: docs.Open(os.path.abspath(dwg_path), True))  # ReadOnly
        # Silenciar el dialogo "Version Conflict" de AEC/proxy objects antes
        # del SaveAs a formato antiguo. PROXYGRAPHICS=1 guarda proxy graphics
        # (los AEC quedan visibles pero no editables, suficiente para croquis).
        for var, val in (("FILEDIA", 0), ("EXPERT", 5),
                         ("PROXYGRAPHICS", 1), ("PROXYNOTICE", 0)):
            try:
                _com_busy(lambda var=var, val=val: doc.SetVariable(var, val), intentos=20)
            except Exception:
                pass
        _com_busy(lambda: doc.SaveAs(os.path.abspath(tmp), 1))  # 1 = acR12_dxf
        return tmp
    except Exception as e:
        sys.exit(
            "ERROR: dwg2dxf (LibreDWG) no esta en PATH y AutoCAD COM fallo:\n"
            f"  {e}\n"
            "  - Instala LibreDWG: https://www.gnu.org/software/libredwg/\n"
            "  - O abre AutoCAD y reintenta (fallback COM usa la instancia activa)\n"
            "  - O exporta el DXF desde AutoCAD (Archivo > Guardar como > DXF)"
        )


def renderizar(dxf_path: str, x: float, y: float, size: float,
               out: str, dpi: int, anclar_esquina: bool) -> None:
    """Renderiza solo la ventana [x0,x1]x[y0,y1] preservando unidades (cm)."""
    media = size / 2.0
    x0, x1 = (x, x + size) if anclar_esquina else (x - media, x + media)
    y0, y1 = (y, y + size) if anclar_esquina else (y - media, y + media)

    doc = ezdxf.readfile(dxf_path)
    msp = doc.modelspace()

    lado_pulg = size / CM_POR_PULGADA
    fig = plt.figure(figsize=(lado_pulg, lado_pulg))
    ax = fig.add_axes([0, 0, 1, 1])
    ax.set_axis_off()

    Frontend(RenderContext(doc), MatplotlibBackend(ax)).draw_layout(msp)

    ax.set_xlim(x0, x1)
    ax.set_ylim(y0, y1)
    ax.set_aspect("equal")
    fig.savefig(out, dpi=dpi)
    plt.close(fig)


def selftest() -> None:
    """Genera un DXF minimo con entidades en coords conocidas, lo renderiza
    y verifica que el PNG no este vacio. Falla si la cadena se rompe."""
    doc = ezdxf.new(setup=True)
    msp = doc.modelspace()
    msp.add_line((90, 90), (110, 110))
    msp.add_circle((100, 100), radius=5)
    doc.layers.add("CROQUIS_TEST")
    tmp = tempfile.NamedTemporaryFile(suffix=".dxf", delete=False).name
    doc.saveas(tmp)
    out = tmp.replace(".dxf", "_selftest.png")
    renderizar(tmp, x=100, y=100, size=10, out=out, dpi=72,
               anclar_esquina=False)
    kb = os.path.getsize(out) / 1024
    assert kb > 0.5, f"PNG vacio ({kb:.1f} KB)"
    os.unlink(tmp)
    os.unlink(out)
    print(f"OK self-test: PNG de {kb:.1f} KB generado correctamente.")


def _leer_xy(ruta: str) -> tuple[float, float] | tuple[None, None]:
    """Lee (x, y) de B1/C1 del archivo (xlsx o csv). (None, None) si invalido.

    CSV: B=columna 1, C=columna 2 de la primera fila no vacia (0-indexed).
    xlsx: B1/C1 de la hoja activa via openpyxl. Mismo contrato que gui_batch.
    """
    if ruta.lower().endswith(".csv"):
        import csv
        try:
            with open(ruta, newline="", encoding="utf-8-sig") as f:
                fila = next(csv.reader(f), None)
            if not fila or len(fila) < 3:
                return None, None
            return float(fila[1]), float(fila[2])
        except (TypeError, ValueError, OSError):
            return None, None
    import openpyxl
    try:
        wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
        try:
            ws = wb.active
            return float(ws["B1"].value), float(ws["C1"].value)
        finally:
            wb.close()
    except (TypeError, ValueError, OSError):
        return None, None


def gui() -> None:
    """Modo interactivo: dialogo para carpeta raiz + Excel con coords X,Y.

    Espera en la carpeta raiz un unico .dwg o .dxf. El ortomosaico de fotos
    (si existe como subcarpeta) se ignora: el croquis sale del CAD, no de las
    fotos. Las coordenadas del punto se leen de B1 (X) y C1 (Y) del Excel que
    el usuario seleccione. size fijo en 100x100 cm; para otro tamano, CLI --size.
    """
    import pathlib
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    carpeta = filedialog.askdirectory(
        title="Seleccione la carpeta raiz (con el .dwg/.dxf)",
        parent=root)
    if not carpeta:
        return

    cad = sorted(pathlib.Path(carpeta).glob("*.dwg")) \
        + sorted(pathlib.Path(carpeta).glob("*.dxf"))
    if not cad:
        messagebox.showerror("Croquis", f"No hay .dwg/.dxf en:\n{carpeta}")
        return
    if len(cad) > 1:
        messagebox.showerror(
            "Croquis",
            f"Hay {len(cad)} archivos CAD en la raiz. Deje solo uno:\n{carpeta}")
        return
    archivo = str(cad[0])

    xlsx = filedialog.askopenfilename(
        title="Seleccione el archivo con las coordenadas (B1=X, C1=Y)",
        filetypes=[("Excel/CSV", "*.xlsx *.csv"), ("Excel", "*.xlsx"),
                   ("CSV", "*.csv")],
        parent=root,
    )
    if not xlsx:
        return
    x, y = _leer_xy(xlsx)
    if x is None:
        messagebox.showerror(
            "Croquis",
            f"No se pudieron leer B1/C1 como numeros en:\n{xlsx}")
        return

    # ponytail: size 100 fijo (usuario pidio 100x100). Para otro tamano, modo CLI --size.
    SIZE_GUI = 100.0
    es_dwg = archivo.lower().endswith(".dwg")
    dxf = dwg_a_dxf(archivo) if es_dwg else archivo
    try:
        out = os.path.splitext(archivo)[0] + "_croquis.png"
        renderizar(dxf, x, y, SIZE_GUI, out, dpi=150, anclar_esquina=False)
    except Exception as e:
        messagebox.showerror("Croquis", f"Error al renderizar:\n{e}")
        return
    finally:
        if es_dwg:
            os.unlink(dxf)
    kb = os.path.getsize(out) / 1024
    messagebox.showinfo(
        "Croquis",
        f"OK: {os.path.basename(out)}\n{kb:.1f} KB - "
        f"ventana 100x100 cm en ({x},{y})\nGuardado en:\n{out}")


def _seleccionar_multiple(titulo, opciones, parent):
    """Dialogo multiselect. Devuelve lista de indices elegidos, o [] si cancela."""
    import tkinter as tk
    top = tk.Toplevel(parent)
    top.title(titulo)
    top.transient(parent)
    lb = tk.Listbox(top, selectmode="multiple",
                    height=min(25, len(opciones)), width=72)
    for op in opciones:
        lb.insert("end", op)
    lb.pack(fill="both", expand=True, padx=8, pady=8)
    sel = []

    def ok():
        sel.extend(lb.curselection())
        top.destroy()
    fr = tk.Frame(top)
    fr.pack(pady=4)
    tk.Button(fr, text="Procesar", command=ok).pack(side="left", padx=8)
    tk.Button(fr, text="Cancelar", command=top.destroy).pack(side="left", padx=8)
    top.grab_set()
    parent.wait_window(top)
    return sel


def gui_batch() -> None:
    """Modo batch: croquis para varios puntos leyendo X,Y de Excels.

    Carpeta raiz: 1 CAD (.dwg/.dxf) + subcarpetas de puntos. Cada subcarpeta:
    1 Excel con nombre ^\\d+_.+_.+\\.xlsx$. Del Excel: B1=X, C1=Y. PNG 100x100
    junto al Excel. Sin limite de puntos (cuello: ~1s/PNG + 1 carga de CAD).
    """
    import re
    from tkinter import filedialog, messagebox
    import openpyxl

    PATRON = re.compile(r'^\d+_.+_.+\.xlsx$', re.IGNORECASE)
    import pathlib
    import tkinter as tk

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)

    carpeta = filedialog.askdirectory(
        title="Carpeta raiz (CAD + subcarpetas de puntos)",
        parent=root)
    if not carpeta:
        return

    cad = sorted(pathlib.Path(carpeta).glob("*.dwg")) \
        + sorted(pathlib.Path(carpeta).glob("*.dxf"))
    if not cad:
        messagebox.showerror("Croquis batch", f"No hay .dwg/.dxf en:\n{carpeta}")
        return
    if len(cad) > 1:
        messagebox.showerror(
            "Croquis batch",
            f"Hay {len(cad)} CAD en la raiz. Deje solo uno.")
        return
    archivo_cad = str(cad[0])

    candidatos = []
    for sub in sorted(p for p in pathlib.Path(carpeta).iterdir() if p.is_dir()):
        xl = sorted(f for f in sub.iterdir() if PATRON.match(f.name))
        if xl:
            candidatos.append((sub, xl[0]))
    if not candidatos:
        messagebox.showerror(
            "Croquis batch",
            "No hay subcarpetas con Excel con el formato #_Nombre_fecha.xlsx")
        return

    etiquetas = [f"{s.name}  <-  {x.name}" for s, x in candidatos]
    idx = _seleccionar_multiple(
        f"Subcarpetas a procesar ({len(candidatos)} con Excel valido)",
        etiquetas, root)
    if not idx:
        return
    elegidos = [candidatos[i] for i in idx]

    SIZE = 100.0
    es_dwg = archivo_cad.lower().endswith(".dwg")
    dxf = dwg_a_dxf(archivo_cad) if es_dwg else archivo_cad
    ok_list, fallos = [], []
    try:
        for sub, xl in elegidos:
            try:
                wb = openpyxl.load_workbook(str(xl), data_only=True,
                                            read_only=True)
                ws = wb.active
                bx, cy = ws['B1'].value, ws['C1'].value
                wb.close()
                if not isinstance(bx, (int, float)) or not isinstance(cy, (int, float)):
                    fallos.append(f"{sub.name}: B1/C1 no numerico ({bx!r},{cy!r})")
                    continue
                out = str(sub / (sub.name + "_croquis.png"))
                renderizar(dxf, bx, cy, SIZE, out, dpi=150,
                           anclar_esquina=False)
                kb = pathlib.Path(out).stat().st_size / 1024
                ok_list.append(f"{sub.name}: {kb:.1f}KB en ({bx},{cy})")
            except Exception as e:
                fallos.append(f"{sub.name}: {e}")
    finally:
        if es_dwg:
            os.unlink(dxf)

    resumen = f"OK {len(ok_list)}/{len(elegidos)}\n\n" + "\n".join(ok_list)
    if fallos:
        resumen += "\n\nFALLOS:\n" + "\n".join(fallos)
    messagebox.showinfo("Croquis batch", resumen)


def main() -> None:
    p = argparse.ArgumentParser(description="DWG/DXF -> PNG croquis.")
    p.add_argument("--input", help="Ruta .dxf o .dwg")
    p.add_argument("--x", type=float, help="Coordenada X del punto (cm)")
    p.add_argument("--y", type=float, help="Coordenada Y del punto (cm)")
    p.add_argument("--size", type=float, default=100.0,
                   help="Lado de la ventana en cm (defecto: 100)")
    p.add_argument("--out", default="croquis.png", help="PNG de salida")
    p.add_argument("--dpi", type=int, default=150, help="DPI (defecto: 150)")
    p.add_argument("--esquina", action="store_true",
                   help="Anclar esquina SW en (X,Y) en vez de centrar")
    p.add_argument("--self-test", action="store_true",
                   help="Verifica la cadena de render y sale")
    p.add_argument("--gui", action="store_true",
                   help="Modo interactivo: dialogo para carpeta raiz + coords X,Y")
    p.add_argument("--batch", action="store_true",
                   help="Modo batch: lee X,Y de Excels en subcarpetas (B1,C1)")
    a = p.parse_args()

    if len(sys.argv) == 1:
        gui()
        return
    if a.self_test:
        selftest()
        return
    if a.gui:
        gui()
        return
    if a.batch:
        gui_batch()
        return
    for req, nombre in ((a.input, "--input"), (a.x, "--x"), (a.y, "--y")):
        if req is None:
            sys.exit(f"Falta {nombre}. Usa --self-test para verificar.")
    es_dwg = a.input.lower().endswith(".dwg")
    dxf = dwg_a_dxf(a.input) if es_dwg else a.input
    try:
        renderizar(dxf, a.x, a.y, a.size, a.out, a.dpi, a.esquina)
    finally:
        if es_dwg:
            os.unlink(dxf)
    kb = os.path.getsize(a.out) / 1024
    print(f"OK: {a.out} ({kb:.1f} KB)  ventana {a.size}x{a.size} cm en "
          f"({a.x},{a.y})")


if __name__ == "__main__":
    main()
