"""Croquis de localizacion via AutoCAD/Civil 3D COM.

Captura un area cuadrada centrada en (x, y) del DWG usando la instancia de
Civil 3D ya abierta (pywin32). Exporta por captura de pantalla (ImageGrab), no
por PNGOUT (este abre dialogos modales al renderizar ortomosaicos ECW).

Flujo principal (--batch):
    elegir carpeta RAIZ (con DWG + Ortomosaico + subcarpetas de puntos)
    -> por cada punto: centro (X,Y) de B1 y C1 del primer .xlsx de su carpeta
       (formato #_nombre_fecha.xlsx, sin encabezado)
    -> Civil 3D: abrir DWG -> esperar carga ECW -> CLEANSCREENON -> ZoomWindow
       -> esperar idle -> capturar area cliente -> recortar -> PNG en la
       carpeta del punto (mismo nombre que la subcarpeta)

Dependencias:  pywin32, Pillow  (pip install pywin32 Pillow)
Requiere:      Civil 3D / AutoCAD corriendo en esta maquina (Windows).
Config (env):  CROQUIS_SIZE (lado de la ventana en unidades del DWG, def 200),
               CROQUIS_PROGID (def "AutoCAD.Application"),
               CROQUIS_CROP_TOP (px a recortar arriba, def 60),
               CROQUIS_KEEP_OPEN=1 (no cerrar el DWG entre capturas).
"""

import csv
import os
import sys
import time
from pathlib import Path

import openpyxl
import pythoncom
import pywintypes
import win32com.client
import win32con
import win32gui
from PIL import ImageGrab

# hresult que indican "Civil 3D ocupado" -> reintentar.
_BUSY = {-2147418111, -2147418102}  # RPC_E_CALL_REJECTED, RPC_E_SERVERCALL_RETRYLATER


def _com(fn, intentos=240, espera=1.0):
    """Ejecuta una llamada COM reintentando si Civil 3D esta ocupado
    (arranque en frio / carga de ECW de minutos). Budget ~4 min por defecto."""
    for k in range(intentos):
        try:
            return fn()
        except pywintypes.com_error as e:
            if getattr(e, "hresult", None) in _BUSY and k + 1 < intentos:
                time.sleep(espera)
                continue
            raise
    raise RuntimeError(
        "Civil 3D esta bloqueado por un dialogo modal (probablemente la ventana "
        "del ortomosaico). Cancela/cierra ese cuadro en Civil 3D y vuelve a ejecutar."
    )


def _pto(x: float, y: float, z: float = 0.0):
    """Punto 3D como VARIANT(VT_ARRAY|VT_R8) que exige la API COM de AutoCAD."""
    return win32com.client.VARIANT(pythoncom.VT_ARRAY | pythoncom.VT_R8, (x, y, z))


def _asegurar_visible(acad):
    """Fuerza visible la ventana principal. ZoomWindow la exige.
    ponytail: el setter acad.Visible falla en despacho dinamico de Civil 3D 2026;
    mostramos via HWND (API Win32) que no depende de la type-lib."""
    try:
        if acad.Visible:
            return
    except Exception:
        pass
    try:
        hwnd = int(acad.HWND)
        win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
        win32gui.ShowWindow(hwnd, win32con.SW_SHOW)
        win32gui.SetForegroundWindow(hwnd)
        time.sleep(0.5)
    except Exception:
        pass


def conectar_autocad(prog_id: str | None = None):
    """Se adhiere a la instancia de AutoCAD/Civil 3D abierta (early-bound).
    Early binding resuelve Open/Count/Item/GetVariable que el despacho dinamico rompe.
    Envuelto en _com: si Civil 3D esta ocupado al conectar, reintenta."""
    pid = prog_id or os.environ.get("CROQUIS_PROGID", "AutoCAD.Application")
    from win32com.client import gencache

    def _ensure():
        try:
            return gencache.EnsureDispatch(pid)
        except pywintypes.com_error:
            raise  # dejar que _com reintente si es busy
        except Exception:
            return win32com.client.Dispatch(pid)

    acad = _com(_ensure)
    _asegurar_visible(acad)
    return acad


def _maximizar(acad):
    """Maximiza la ventana de Civil 3D. Llamar ANTES del ZoomWindow: maximizar
    despues hace que Civil 3D reajuste la vista a la extension completa."""
    try:
        hwnd = int(acad.HWND)
        win32gui.ShowWindow(hwnd, win32con.SW_RESTORE)
        win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
        win32gui.SetForegroundWindow(hwnd)
        time.sleep(0.5)
    except Exception:
        pass


def _esperar_idle(acad, budget_s=300):
    """Espera a que Civil 3D deje de estar ocupado (carga de ECW completada).
    Durante la carga las llamadas COM devuelven CALL_REJECTED; sondeamos hasta
    que una llamada barata (Version) ceda. Devuelve True si quedo idle."""
    deadline = time.time() + budget_s
    while time.time() < deadline:
        try:
            _ = acad.Version
            return True
        except pywintypes.com_error as e:
            if getattr(e, "hresult", None) in _BUSY:
                time.sleep(1.0)
                continue
            return True
    return False


def _obtener_doc(acad, path: str):
    """Activa el DWG si ya esta abierto; si no, lo abre ReadOnly.
    Reabrir un doc abierto dispara un dialogo modal que cuelga el COM."""
    nombre = os.path.basename(path).lower()
    docs = acad.Documents
    try:
        for i in range(_com(lambda: docs.Count)):
            d = _com(lambda: docs.Item(i))
            if os.path.basename(str(d.FullName)).lower() == nombre:
                _com(lambda: d.Activate())
                return d
    except Exception:
        pass
    _com(lambda: docs.Open(os.path.abspath(path), True))  # ReadOnly
    return acad.ActiveDocument


def _anadir_support_path(acad, carpeta):
    """Anade la carpeta de referencias al Support File Search Path.

    AutoCAD resuelve XREFs/imagenes cuyo path guardado no existe buscando en el
    support path; esto evita el dialogo 'referencia no resuelta' que cuelga el Open.
    """
    if not carpeta:
        return False
    try:
        files = acad.Preferences.Files
        actual = str(files.SupportPath or "")
        partes = [p for p in actual.split(";") if p]
        carpeta_abs = os.path.abspath(carpeta)
        if carpeta_abs not in partes:
            partes.append(carpeta_abs)
            files.SupportPath = ";".join(partes)
        return True
    except Exception as e:
        print(f"  (SupportPath no seteable: {e})", flush=True)
        return False


def _interact_gui():
    """Popups para elegir DWG + carpeta de referencias + centro (x,y).
    Devuelve (dwg, refs, x, y, salida) o None si el usuario cancela."""
    import tkinter as tk
    from tkinter import filedialog, simpledialog

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    dwg = filedialog.askopenfilename(
        title="Selecciona el DWG",
        filetypes=[("DWG", "*.dwg"), ("DXF", "*.dxf"), ("Todos", "*.*")],
    )
    if not dwg:
        return None
    refs = filedialog.askdirectory(
        title="Carpeta de referencias (Ortomosaico) - Cancelar si no aplica"
    )
    x = simpledialog.askfloat("Centro X", "Coordenada X del centro del croquis:")
    if x is None:
        return None
    y = simpledialog.askfloat("Centro Y", "Coordenada Y del centro del croquis:")
    if y is None:
        return None
    salida = filedialog.asksaveasfilename(
        title="Guardar PNG como",
        defaultextension=".png",
        filetypes=[("PNG", "*.png")],
        initialfile=str(Path(dwg).with_suffix(".png").name),
    ) or str(Path(dwg).with_suffix(".png"))
    root.destroy()
    return dwg, (refs or None), x, y, salida


def _minimizar_consola():
    """Minimiza la ventana de la consola para que no salga en la captura.
    Devuelve el HWND de la consola (0 si no hay)."""
    import ctypes
    hwnd = ctypes.windll.kernel32.GetConsoleWindow()
    if hwnd:
        ctypes.windll.user32.ShowWindow(hwnd, 6)  # SW_MINIMIZE
    return hwnd


def _restaurar_consola(hwnd):
    """Restaura la consola minimizada por _minimizar_consola."""
    if hwnd:
        import ctypes
        ctypes.windll.user32.ShowWindow(hwnd, 9)  # SW_RESTORE


def _forzar_frente(hwnd):
    """Trae el HWND de AutoCAD al frente superando las restricciones de
    SetForegroundWindow (Windows bloquea el robo de foco entre procesos).
    Usa AttachThreadInput para sincronizar el hilo foreground con el nuestro.
    Devuelve True si AutoCAD quedo como ventana foreground."""
    import ctypes
    user32 = ctypes.windll.user32
    kernel32 = ctypes.windll.kernel32
    if user32.GetForegroundWindow() == hwnd:
        return True
    fg = user32.GetForegroundWindow()
    fg_tid = user32.GetWindowThreadProcessId(fg, None)
    our_tid = kernel32.GetCurrentThreadId()
    attached = user32.AttachThreadInput(our_tid, fg_tid, True)
    try:
        user32.ShowWindow(hwnd, 9)  # SW_RESTORE (por si esta minimizada)
        user32.SetForegroundWindow(hwnd)
        user32.BringWindowToTop(hwnd)
    finally:
        if attached:
            user32.AttachThreadInput(our_tid, fg_tid, False)
    time.sleep(0.3)
    return user32.GetForegroundWindow() == hwnd


def _ventanas_autocad_visibles():
    """Lista (hwnd, titulo) de ventanas visibles de AutoCAD/Civil 3D.
    Para el aviso cuando hay varias instancias o ninguna al frente."""
    import ctypes
    user32 = ctypes.windll.user32
    found = []
    EnumWindowsProc = ctypes.WINFUNCTYPE(ctypes.c_bool, ctypes.c_void_p, ctypes.c_void_p)

    def _cb(hwnd, _lparam):
        if not user32.IsWindowVisible(hwnd):
            return True
        length = user32.GetWindowTextLengthW(hwnd)
        if length == 0:
            return True
        buf = ctypes.create_unicode_buffer(length + 1)
        user32.GetWindowTextW(hwnd, buf)
        titulo = buf.value
        if any(k in titulo for k in ("AutoCAD", "Civil 3D", "Autodesk")):
            found.append((int(hwnd), titulo))
        return True
    user32.EnumWindows(EnumWindowsProc(_cb), 0)
    return found


def _capturar_pantalla(acad, salida_png):
    """Captura el area cliente de Civil 3D. De usarse tras CLEANSCREENON y
    maximizar, el area cliente es (casi todo) el lienzo de dibujo.
    Minimiza la consola durante la captura para que no aparezca en la imagen."""
    consola = _minimizar_consola()
    try:
        hwnd = int(_com(lambda: acad.HWND, intentos=300))
        if not _forzar_frente(hwnd):
            visibles = _ventanas_autocad_visibles()
            print("  ADVERTENCIA: AutoCAD no quedo al frente; la captura puede no ser de AutoCAD.", flush=True)
            if visibles:
                print("  Ventanas de AutoCAD detectadas:", flush=True)
                for h, t in visibles:
                    print(f"    - [{h}] {t}", flush=True)
        if os.environ.get("CROQUIS_CONFIRMAR_VENTANA"):
            input("  ENTER para capturar (verifica que AutoCAD este al frente)... ")
        time.sleep(0.5)
        cl = win32gui.GetClientRect(hwnd)
        p1 = win32gui.ClientToScreen(hwnd, (0, 0))
        p2 = win32gui.ClientToScreen(hwnd, (cl[2], cl[3]))
        bbox = (p1[0], p1[1], p2[0], p2[1])
        img = ImageGrab.grab(bbox, all_screens=True)
        # Recortar la franja superior (menu/barra de herramientas que CleanScreen no oculta).
        crop_top = int(os.environ.get("CROQUIS_CROP_TOP", "60"))
        if crop_top > 0:
            img = img.crop((0, crop_top, img.width, img.height))
        print(f"  imagen {img.size} {img.mode} bbox={bbox} crop_top={crop_top}", flush=True)
        out = os.path.abspath(salida_png)
        img.save(out)
        return os.path.getsize(out) // 1024
    finally:
        _restaurar_consola(consola)


def _volcar_rutas_imagenes(doc, archivo_salida):
    """Diagnostico: vuelca referencias externas del DWG.
    1) imagenes (ACAD_IMAGE_DICT)  2) XREFs (tabla BLOCK con path)."""
    out = os.path.abspath(archivo_salida).replace("\\", "/")
    lisp = (
        "(progn"
        '(setq f (open "' + out + '" "w"))'
        '(write-line "=== IMAGENES (ACAD_IMAGE_DICT) ===" f)'
        '(setq de (cdr (assoc -1 (dictsearch (namedobjdict) "ACAD_IMAGE_DICT"))))'
        "(if de (progn"
        "(setq e (dictnext de T))"
        "(while e"
        "(setq p (cdr (assoc 1 (entget (cdr (assoc -1 e))))))"
        '(write-line p f)'
        "(setq e (dictnext de))"
        ")))"
        '(write-line "=== XREFS (tabla BLOCK) ===" f)'
        "(setq b (tblnext \"BLOCK\" T))"
        "(while b"
        "(if (assoc 1 b) (write-line (strcat (cdr (assoc 2 b)) \" => \" (cdr (assoc 1 b))) f))"
        "(setq b (tblnext \"BLOCK\"))"
        ")"
        "(close f)"
        ")"
    )
    _com(lambda: doc.SendCommand(lisp + "\n"))
    time.sleep(0.6)


def capturar_croquis(
    archivo_dwg: str,
    x: float,
    y: float,
    salida_png: str,
    size_cm: float | None = None,
    acad=None,
    refs_folder: str | None = None,
) -> tuple[str, int]:
    """Abre el DWG, recorta size x size centrado en (x,y) y captura PNG (ImageGrab).

    Devuelve (ruta_png, kilobytes). Lanza si Civil 3D reporta error.
    ponytail: captura de pantalla = resolucion del monitor; para DPI fijo usar
    PLOT a configuracion PNG (mas codigo) si la salida sale borrosa.
    """
    size = float(size_cm or os.environ.get("CROQUIS_SIZE", 200))
    half = size / 2.0
    acad = acad or conectar_autocad()
    if refs_folder:
        _anadir_support_path(acad, refs_folder)
    doc = None
    filedia_prev = None
    try:
        doc = _obtener_doc(acad, archivo_dwg)
        # Esperar a que termine la carga del DWG + ortomosaico ANTES de tocar el
        # doc (sino doc.Name recibe RPC_E_CALL_REJECTED mientras AutoCAD carga).
        print("  esperando idle (carga del DWG/ortomosaico)...", flush=True)
        _esperar_idle(acad)
        print(f"  doc: {_com(lambda: doc.Name)}", flush=True)
        # FILEDIA=0 antes de cualquier comando (evita dialogos que cuelgan COM).
        try:
            filedia_prev = _com(lambda: doc.GetVariable("FILEDIA"))
            _com(lambda: doc.SetVariable("FILEDIA", 0))
            print("  FILEDIA=0 via API", flush=True)
        except Exception:
            _com(lambda: doc.SendCommand('(setvar "FILEDIA" 0)\n'))
            time.sleep(0.3)
            print("  FILEDIA=0 via SendCommand", flush=True)
        # Maximizar ANTES del ZoomWindow (maximizar despues reajusta la vista).
        _maximizar(acad)
        # CleanScreen: oculta ribbon y paletas (incluida Referencias externas).
        try:
            _com(lambda: doc.SendCommand('_.CLEANSCREENON\n'))
            time.sleep(0.5)
        except Exception:
            pass
        # Ventana cuadrada centrada en (x, y), plano Z=0.
        _com(lambda: acad.ZoomWindow(_pto(x - half, y - half), _pto(x + half, y + half)))
        print("  ZoomWindow OK", flush=True)
        # Esperar a que carguen los tiles del ECW de la nueva vista + pintado final.
        _esperar_idle(acad)
        time.sleep(float(os.environ.get("CROQUIS_WAIT_AFTER_ZOOM", "8")))
        kb = _capturar_pantalla(acad, salida_png)
        return os.path.abspath(salida_png), kb
    finally:
        try:
            if doc is not None:
                try:
                    _com(lambda: doc.SendCommand('_.CLEANSCREENOFF\n'), intentos=10)
                except Exception:
                    pass
                if filedia_prev is not None:
                    _com(lambda: doc.SetVariable("FILEDIA", filedia_prev), intentos=10)
                if not os.environ.get("CROQUIS_KEEP_OPEN"):
                    _com(lambda: doc.Close(False), intentos=10)
        except Exception:
            pass


def resolver_coords(supabase_url: str, supabase_key: str, codigo: int):
    """(x, y) del punto desde Supabase por numero_serie (join a coordenadas_gps).

    Devuelve (coordenada_x, coordenada_y) en unidades del dibujo o None.
    """
    url = (
        f"{supabase_url.rstrip('/')}/rest/v1/puntos_ferroviarios"
        f"?select=numero_serie,coordenadas_gps(coordenada_x,coordenada_y)"
        f"&numero_serie=eq.{int(codigo)}&limit=1"
    )
    req = urllib.request.Request(
        url,
        headers={
            "apikey": supabase_key,
            "Authorization": f"Bearer {supabase_key}",
            "Accept": "application/json",
        },
    )
    with urllib.request.urlopen(req, timeout=10) as r:
        data = json.load(r)
    if not data:
        return None
    coords = data[0].get("coordenadas_gps") or []
    if not coords:
        return None
    c = coords[0]
    return float(c["coordenada_x"]), float(c["coordenada_y"])


def extraer_codigo(ruta: str) -> int | None:
    """Primer entero del nombre de archivo o carpeta -> numero_serie.
    ponytail: heuristica de naming; ajustar el regex a la convencion del NAS.
    """
    m = re.search(r"\d+", Path(ruta).name)
    return int(m.group()) if m else None


def _demo():
    """Self-check: conecta a AutoCAD (reporta version) y resuelve coords si hay env."""
    pythoncom.CoInitialize()
    print("[demo] Conectando a AutoCAD...")
    try:
        acad = conectar_autocad()
        print(f"  OK: {acad.Version}")
    except Exception as e:
        print(f"  FAIL: {e}")
        return 1
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_ANON_KEY")
    if url and key:
        cod = int(os.environ.get("CROQUIS_DEMO_CODIGO", "1"))
        print(f"[demo] Resolviendo coords para numero_serie={cod}...")
        xy = resolver_coords(url, key, cod)
        print(f"  -> {xy}" if xy else "  -> no encontrado")
    return 0


def _leer_centro(carpeta_punto):
    """Centro (x, y) del PRIMER punto del primer .xlsx o .csv de la carpeta.

    xlsx: B1/C1 (o B2/C2 si la 1 es encabezado). CSV: cols 1,2 de la 1a fila
    (0-indexed; mismo contrato que _leer_xy de dwg-to-croquis.py). None si no
    hay coords numericas. X en col B, Y en col C."""
    archivos = sorted(f for f in os.listdir(carpeta_punto)
                      if f.lower().endswith((".xlsx", ".csv")))
    if not archivos:
        return None
    p = os.path.join(carpeta_punto, archivos[0])
    if p.lower().endswith(".csv"):
        import csv
        try:
            with open(p, newline="", encoding="utf-8-sig") as f:
                filas = list(csv.reader(f))
            for fila in filas[:2]:  # B1/C1 y B2/C2 equivalentes
                if len(fila) >= 3:
                    try:
                        return float(fila[1]), float(fila[2])
                    except (TypeError, ValueError):
                        continue
        except OSError:
            pass
        return None
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    try:
        ws = wb.active
        for x_ref, y_ref in (("B1", "C1"), ("B2", "C2")):
            try:
                return float(ws[x_ref].value), float(ws[y_ref].value)
            except (TypeError, ValueError):
                continue
        return None
    finally:
        wb.close()


def _leer_puntos(carpeta_punto, max_n=0):
    """Lista de (x, y, etiqueta) del primer .xlsx o .csv de la carpeta.

    Recorre TODAS las filas con coords B/C numericas (no solo la primera).
    Etiqueta = "p{N}" (1-based). max_n=0 = sin tope. Mismo parseo que
    _leer_centro pero multiple. Vacio si no hay archivo o no hay filas validas.
    """
    archivos = sorted(f for f in os.listdir(carpeta_punto)
                      if f.lower().endswith((".xlsx", ".csv")))
    if not archivos:
        return []
    p = os.path.join(carpeta_punto, archivos[0])
    out = []
    if p.lower().endswith(".csv"):
        import csv
        try:
            with open(p, newline="", encoding="utf-8-sig") as f:
                filas = list(csv.reader(f))
        except OSError:
            return []
        for i, fila in enumerate(filas, 1):
            if len(fila) >= 3:
                try:
                    out.append((float(fila[1]), float(fila[2]), f"p{i}"))
                except (TypeError, ValueError):
                    continue
            if max_n and len(out) >= max_n:
                break
        return out
    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    try:
        ws = wb.active
        r, sin_data = 1, 0
        while sin_data < 5 and (not max_n or len(out) < max_n):
            try:
                x = float(ws[f"B{r}"].value)
                y = float(ws[f"C{r}"].value)
                out.append((x, y, f"p{r}"))
                sin_data = 0
            except (TypeError, ValueError):
                sin_data += 1
            r += 1
            if r > 1000:
                break
        return out
    finally:
        wb.close()


def _elegir_puntos(raiz):
    """Lista de subcarpetas de punto elegidas (multi-seleccion)."""
    import tkinter as tk

    subdirs = sorted(
        d for d in os.listdir(raiz)
        if os.path.isdir(os.path.join(raiz, d)) and not d.startswith(".")
    )
    if not subdirs:
        return []
    top = tk.Tk()
    top.title("3/3 Selecciona carpeta(s) de punto")
    top.attributes("-topmost", True)
    tk.Label(top, text=f"Subcarpetas de:\n{raiz}").pack(anchor="w", padx=8, pady=4)
    lb = tk.Listbox(top, selectmode="multiple", width=64, height=min(24, len(subdirs)))
    for d in subdirs:
        lb.insert("end", d)
    lb.pack(fill="x", padx=8)
    ok = {"v": False}

    def _ok():
        ok["v"] = True
        top.destroy()

    bb = tk.Frame(top)
    bb.pack(pady=6)
    tk.Button(bb, text="Cancelar", command=top.destroy).pack(side="left", padx=4)
    tk.Button(bb, text="Generar croquis", command=_ok).pack(side="left", padx=4)
    top.mainloop()
    if not ok["v"]:
        return []
    return [os.path.join(raiz, lb.get(i)) for i in lb.curselection()]


def _fmt_dur(s: float) -> str:
    """Segundos -> 'Xs' / 'Xm Ys' / 'Xh Ym'. Para la barra de tiempo restante."""
    s = int(s)
    if s < 60:
        return f"{s}s"
    m, sec = divmod(s, 60)
    if m < 60:
        return f"{m}m {sec}s"
    h, m = divmod(m, 60)
    return f"{h}h {m}m"


def _batch_gui():
    """Flujo completo con pickers:
    1) Carpeta RAIZ (con DWG + subcarpetas de puntos)
    2) DWG (auto si hay uno solo)
    3) Carpeta(s) de punto (multi-seleccion)
    4) Ventana de progreso: estado actual + barra + tiempo restante + log en vivo.
    Lee el centro (X,Y) de B1/C1 del .xlsx de cada punto, genera el croquis y
    guarda el PNG en la carpeta del punto (mismo nombre que la subcarpeta)."""
    import tkinter as tk
    from tkinter import filedialog, messagebox

    root = tk.Tk()
    root.withdraw()
    root.attributes("-topmost", True)
    raiz = filedialog.askdirectory(
        title="1/3 Carpeta RAIZ (contiene el DWG y subcarpetas de puntos)",
        parent=root,
    )
    if not raiz:
        return 0

    dwgs = sorted(f for f in os.listdir(raiz) if f.lower().endswith(".dwg"))
    if not dwgs:
        messagebox.showerror("Error", f"No hay archivo .dwg en:\n{raiz}")
        return 0
    if len(dwgs) == 1:
        dwg = os.path.join(raiz, dwgs[0])
    else:
        dwg = filedialog.askopenfilename(
            title="2/3 Elige el DWG a procesar", initialdir=raiz, filetypes=[("DWG", "*.dwg")]
        )
        if not dwg:
            return 0

    refs = os.path.join(raiz, "Ortomosaico")
    if not os.path.isdir(refs):
        refs = None

    puntos = _elegir_puntos(raiz)
    if not puntos:
        return 0

    # Cuantos puntos por subcarpeta (1 = solo B1/C1, 0 = todos los del CSV/xlsx)
    from tkinter import simpledialog
    n_por_sub = simpledialog.askinteger(
        "Puntos por subcarpeta",
        "Cuantos puntos procesar por subcarpeta?\n(0 = todos los del CSV/xlsx)",
        initialvalue=1, minvalue=0, maxvalue=1000, parent=root,
    )
    if n_por_sub is None:
        return 0

    # Carpeta de salida (Cancelar = junto a cada subcarpeta, comportamiento por defecto)
    salida = filedialog.askdirectory(
        title="Carpeta de SALIDA (Cancelar = junto a cada subcarpeta)",
        initialdir=raiz,
        parent=root,
    )
    root.destroy()
    output_dir = salida if salida else None

    return _procesar_batch_gui(dwg, refs, puntos,
                               n_por_sub=n_por_sub, output_dir=output_dir)


def _procesar_batch_gui(dwg, refs, puntos, n_por_sub=1, output_dir=None):
    """Crea la ventana de progreso y procesa los puntos actualizandola.

    Redirige sys.stdout al log widget para que los print(...) del flujo (conectar,
    abrir DWG, leer xlsx, capturar, etc.) aparezcan en vivo sin tocar las
    funciones internas. La barra muestra puntos completados y se estima el tiempo
    restante con el promedio por punto.

    n_por_sub: cuantos puntos (X,Y) procesar por subcarpeta (0 = todos los del
    CSV/xlsx). output_dir: carpeta donde dejar los PNG (None = junto a cada sub).
    Al terminar, abre la carpeta de salida en Explorer."""
    import time as _time
    import tkinter as tk
    from tkinter import ttk, scrolledtext

    # Pre-conteo: total de puntos a procesar (sumando N por subcarpeta)
    plan = []  # [(nombre_sub, [(x, y, label), ...], sub_path), ...]
    for pp in puntos:
        pts = _leer_puntos(pp, n_por_sub)
        plan.append((os.path.basename(pp), pts, pp))
    total = sum(len(pts) for _, pts, _ in plan)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    top = tk.Tk()
    top.title(f"Croquis batch - {total} punto(s) en {len(plan)} subcarpeta(s)")
    top.geometry("760x480")
    # Mantener la ventana SIEMPRE al frente: conectar_autocad/_maximizar hacen
    # SetForegroundWindow sobre AutoCAD y taparian esta ventana si no es topmost.
    top.attributes("-topmost", True)

    estado_var = tk.StringVar(value="Iniciando...")
    tk.Label(top, textvariable=estado_var, font=("Segoe UI", 10, "bold"),
             anchor="w").pack(fill="x", padx=10, pady=(10, 2))

    pb = ttk.Progressbar(top, maximum=max(total, 1), mode="determinate")
    pb.pack(fill="x", padx=10, pady=2)

    cuenta_var = tk.StringVar(value=f"0 / {total} completados")
    tk.Label(top, textvariable=cuenta_var, anchor="w").pack(fill="x", padx=10)
    eta_var = tk.StringVar(value="Tiempo restante: (tras el 1er punto)")
    tk.Label(top, textvariable=eta_var, anchor="w", fg="#555555").pack(
        fill="x", padx=10, pady=(0, 6))

    log = scrolledtext.ScrolledText(top, width=96, height=22, state="disabled",
                                    font=("Consolas", 9))
    log.pack(fill="both", expand=True, padx=10, pady=(0, 8))

    class _Tee:
        """stdout que escribe al log widget y a la consola real."""
        def __init__(self, consola):
            self.cons = consola

        def write(self, s):
            if s and s.strip():
                log.config(state="normal")
                log.insert("end", s if s.endswith("\n") else s + "\n")
                log.see("end")
                log.config(state="disabled")
            if self.cons:
                try:
                    self.cons.write(s)
                    self.cons.flush()
                except Exception:
                    pass
            top.update_idletasks()

        def flush(self):
            if self.cons:
                try:
                    self.cons.flush()
                except Exception:
                    pass

    consola_real = sys.stdout
    sys.stdout = _Tee(consola_real)
    top.update()

    os.environ["CROQUIS_KEEP_OPEN"] = "1"
    t0 = _time.time()
    hechos = 0
    acad = None
    try:
        estado_var.set("Conectando a AutoCAD...")
        top.update()
        acad = conectar_autocad()
        if refs:
            _anadir_support_path(acad, refs)
        for nombre, pts, pp in plan:
            if not pts:
                print(f"[skip] {nombre}: sin puntos validos en CSV/xlsx", flush=True)
                continue
            for (x, y, label) in pts:
                estado_var.set(f"[{hechos+1}/{total}] {nombre} {label}: capturando ({x}, {y})...")
                cuenta_var.set(f"{hechos} / {total} completados")
                top.update()
                if output_dir:
                    out = os.path.join(output_dir, f"{nombre}_{label}.png")
                else:
                    out = os.path.join(pp, f"{nombre}_{label}.png")
                print(f"[batch] {nombre} {label}  centro=({x}, {y})  -> {out}", flush=True)
                try:
                    o, kb = capturar_croquis(dwg, x, y, out, acad=acad, refs_folder=refs)
                    print(f"  OK: {o} ({kb} KB)", flush=True)
                    hechos += 1
                except Exception as e:
                    print(f"  FAIL: {e}", flush=True)
                pb["value"] = hechos
                cuenta_var.set(f"{hechos} / {total} completados")
                if hechos > 0:
                    avg = (_time.time() - t0) / hechos
                    eta_var.set(f"Tiempo restante aprox.: {_fmt_dur(avg * (total - hechos))}")
                top.update()
    finally:
        sys.stdout = consola_real

    dur = _time.time() - t0
    estado_var.set(f"Listo. {hechos}/{total} capturas OK.")
    eta_var.set(f"Tiempo total: {_fmt_dur(dur)}")
    print(f"\n==== RESUMEN: {hechos}/{total} capturas OK en {_fmt_dur(dur)} ====", flush=True)
    salida_final = output_dir or (puntos[0] if puntos else "")
    if salida_final and os.path.isdir(salida_final):
        try:
            os.startfile(salida_final)
            print(f"Carpeta de salida abierta: {salida_final}", flush=True)
        except Exception as e:
            print(f"(no se pudo abrir la carpeta: {e})", flush=True)
    fr = tk.Frame(top)
    fr.pack(pady=4)
    tk.Button(fr, text="Abrir carpeta de salida",
              command=lambda: os.startfile(salida_final)).pack(side="left", padx=4)
    tk.Button(fr, text="Cerrar", command=top.destroy).pack(side="left", padx=4)
    top.mainloop()

    if acad is not None:
        try:
            _com(lambda: acad.ActiveDocument.Close(False), intentos=10)
        except Exception:
            pass
    return 0


def main(argv):
    """Uso:
      python croquis_com.py --batch                       # pickers: Raiz + DWG + punto(s)
      python croquis_com.py                               # GUI: elige DWG + refs + centro
      python croquis_com.py --gui                         # idem
      python croquis_com.py <dwg> --xy <x> <y> [--refs <carpeta>] [salida]
      python croquis_com.py <dwg> <codigo> [--refs <carpeta>] [salida]
      python croquis_com.py --diag <dwg> [--refs <carpeta>]   # volcar referencias
    """
    pythoncom.CoInitialize()
    gui = False
    diag = False
    batch = False
    refs = None
    xy = None
    posicionales = []
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == "--gui":
            gui = True
            i += 1
        elif a == "--batch":
            batch = True
            i += 1
        elif a == "--diag":
            diag = True
            i += 1
        elif a == "--refs" and i + 1 < len(argv):
            refs = argv[i + 1]
            i += 2
        elif a == "--xy" and i + 2 < len(argv):
            xy = (float(argv[i + 1]), float(argv[i + 2]))
            i += 3
        else:
            posicionales.append(a)
            i += 1

    if batch:
        return _batch_gui()
    if not argv:
        # Doble-clic / Run sin args: arrancar el flujo batch completo
        # (pickers + Civil 3D), no _demo. Mismo patron que dwg-to-croquis.py:287.
        return _batch_gui()
    if diag:
        dwg = posicionales[0] if posicionales else None
        if not dwg:
            print("Uso: croquis_com.py --diag <dwg> [--refs <carpeta>]")
            return 2
        acad = conectar_autocad()
        if refs:
            _anadir_support_path(acad, refs)
        doc = _obtener_doc(acad, dwg)
        dump = os.path.abspath(str(Path(dwg).with_suffix(".imgpaths.txt")))
        _volcar_rutas_imagenes(doc, dump)
        print(f"Rutas de imagen -> {dump}", flush=True)
        try:
            _com(lambda: doc.Close(False), intentos=10)
        except Exception:
            pass
        return 0
    if gui or not posicionales:
        elegido = _interact_gui()
        if not elegido:
            return 0
        dwg, refs_gui, x, y, salida = elegido
        refs = refs or refs_gui
        print(f"[gui] Capturando {dwg} en ({x}, {y}) refs={refs or '(ninguna)'}...", flush=True)
        out, kb = capturar_croquis(dwg, x, y, salida, refs_folder=refs)
        print(f"OK: {out} ({kb} KB)")
        return 0

    dwg = posicionales[0]
    if xy:
        x, y = xy
        salida = posicionales[1] if len(posicionales) > 1 else str(Path(dwg).with_suffix(".png"))
        print(f"[xy] Capturando {dwg} en ({x}, {y}) refs={refs or '(ninguna)'}...", flush=True)
        out, kb = capturar_croquis(dwg, x, y, salida, refs_folder=refs)
        print(f"OK: {out} ({kb} KB)")
        return 0

    codigo = int(posicionales[1]) if len(posicionales) > 1 else extraer_codigo(dwg)
    salida = posicionales[2] if len(posicionales) > 2 else str(Path(dwg).with_suffix(".png"))
    url = os.environ.get("SUPABASE_URL")
    key = os.environ.get("SUPABASE_ANON_KEY")
    if not (url and key):
        print("Falta SUPABASE_URL / SUPABASE_ANON_KEY en el entorno.")
        return 2
    if codigo is None:
        print(f"No se pudo extraer codigo de: {dwg}")
        return 2
    xy = resolver_coords(url, key, codigo)
    if not xy:
        print(f"Punto numero_serie={codigo} sin coordenadas en Supabase.")
        return 3
    x, y = xy
    print(f"Capturando {dwg} en ({x}, {y}) refs={refs or '(ninguna)'}...", flush=True)
    out, kb = capturar_croquis(dwg, x, y, salida, refs_folder=refs)
    print(f"OK: {out} ({kb} KB)")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main(sys.argv[1:]))
    except (RuntimeError, pywintypes.com_error) as e:
        print(f"\n[ERROR] {e}")
        sys.exit(1)
